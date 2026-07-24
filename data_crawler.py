"""Collect model inputs from official online sources and refresh the forecast.

The crawler uses machine-readable Bank of Tanzania spreadsheets for inflation,
policy rate, broad money, and Brent oil. USD/TZS is calculated as the monthly
average of the official daily mean exchange rate.

Only complete months present in every source are appended to the project CSV.
Missing values are reported and never estimated or silently filled.
"""

from __future__ import annotations

import json
from collections import defaultdict
from datetime import datetime
from io import BytesIO
from urllib.parse import urljoin

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

from project_config import INPUT_CSV, SOURCE_LOG_PATH

DATASET_PATH = INPUT_CSV

BOT_ROOT = "https://www.bot.go.tz"
BOT_MONTHLY_STATS_URL = f"{BOT_ROOT}/Publications/Filter/13"
BOT_EXCHANGE_URL = f"{BOT_ROOT}/ExchangeRate/previous_rates"

SOURCE_TITLES = {
    "inflation_pct": "Monthly - NCPI Annual Change",
    "bot_policy_rate_pct": "Monthly - Interest Rates",
    "m2_bn_tzs": "Monthly - Depository Corporations Survey",
    "brent_oil_usd": "Monthly - World Commodity Prices",
}

ROW_MATCHERS = {
    "inflation_pct": lambda label: "headline inflation" in label,
    "bot_policy_rate_pct": lambda label: label == "policy rate",
    "m2_bn_tzs": lambda label: label == "broad money (m2)",
    "brent_oil_usd": lambda label: label == "crude oil (brent)",
}

USER_AGENT = "ProciForecastDashboard/1.0 (official-statistics collector)"


def _month_start(value: datetime) -> pd.Timestamp:
    return pd.Timestamp(value.year, value.month, 1)


def _request_session() -> requests.Session:
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    return session


def _fetch_monthly_spreadsheet_links(session: requests.Session) -> dict[str, str]:
    response = session.get(BOT_MONTHLY_STATS_URL, timeout=45)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, "html.parser")

    links: dict[str, str] = {}
    for anchor in soup.select("a[download][href]"):
        title = " ".join(anchor.get("download", "").split())
        for field, expected_title in SOURCE_TITLES.items():
            if title.casefold() == expected_title.casefold():
                links[field] = urljoin(BOT_ROOT, anchor["href"])

    missing = sorted(set(SOURCE_TITLES) - set(links))
    if missing:
        raise RuntimeError(
            "Bank of Tanzania monthly statistics page is missing required downloads: "
            + ", ".join(missing)
        )
    return links


def _download_workbook(session: requests.Session, url: str):
    response = session.get(url, timeout=60)
    response.raise_for_status()
    if not response.content.startswith(b"PK"):
        raise RuntimeError(f"Expected an XLSX workbook from {url}")
    return load_workbook(BytesIO(response.content), read_only=True, data_only=True)


def _extract_workbook_series(workbook, field: str) -> dict[pd.Timestamp, float]:
    matcher = ROW_MATCHERS[field]
    best_series: dict[pd.Timestamp, float] = {}

    for worksheet in workbook.worksheets:
        rows = list(worksheet.iter_rows(values_only=True))
        date_row = None

        for row in rows:
            if sum(isinstance(value, datetime) for value in row) >= 3:
                date_row = row

            label = next(
                (
                    " ".join(value.split()).casefold()
                    for value in row
                    if isinstance(value, str) and value.strip()
                ),
                "",
            )
            if not date_row or not matcher(label):
                continue

            series = {
                _month_start(date): float(value)
                for date, value in zip(date_row, row)
                if isinstance(date, datetime)
                and isinstance(value, (int, float))
                and not isinstance(value, bool)
            }
            if series and (not best_series or max(series) > max(best_series)):
                best_series = series

    if not best_series:
        raise RuntimeError(f"Could not find {field} in the official BOT workbook.")
    return best_series


def _fetch_exchange_rate_series(
    session: requests.Session,
    start_month: pd.Timestamp,
    end_month: pd.Timestamp,
) -> dict[pd.Timestamp, float]:
    landing = session.get(BOT_EXCHANGE_URL, timeout=45)
    landing.raise_for_status()
    soup = BeautifulSoup(landing.text, "html.parser")
    token = soup.select_one('input[name="__RequestVerificationToken"]')
    if token is None or not token.get("value"):
        raise RuntimeError("Could not obtain the BOT exchange-rate request token.")

    end_date = end_month + pd.offsets.MonthEnd(0)
    response = session.post(
        BOT_EXCHANGE_URL,
        data={
            "__RequestVerificationToken": token["value"],
            "dateFrom": start_month.strftime("%m/%d/%Y"),
            "dateTo": end_date.strftime("%m/%d/%Y"),
        },
        timeout=150,
    )
    response.raise_for_status()

    daily_values: dict[pd.Timestamp, list[float]] = defaultdict(list)
    soup = BeautifulSoup(response.text, "html.parser")
    for row in soup.select("table tr"):
        cells = [cell.get_text(" ", strip=True) for cell in row.find_all(["th", "td"])]
        if len(cells) < 6 or cells[1].upper() != "USD":
            continue
        try:
            value = float(cells[4].replace(",", ""))
            date = pd.to_datetime(cells[5], format="%d-%b-%y")
        except (TypeError, ValueError):
            continue
        daily_values[_month_start(date)].append(value)

    monthly = {
        month: float(sum(values) / len(values))
        for month, values in daily_values.items()
        if values
    }
    if not monthly:
        raise RuntimeError("The BOT exchange-rate page returned no USD observations.")
    return monthly


def _validate_row(row: dict[str, float | str]) -> None:
    checks = {
        "inflation_pct": (-20, 100),
        "usd_tzs_rate": (100, 10000),
        "brent_oil_usd": (1, 1000),
        "m2_bn_tzs": (1, float("inf")),
        "bot_policy_rate_pct": (0, 100),
    }
    for field, (lower, upper) in checks.items():
        value = float(row[field])
        if not lower <= value <= upper:
            raise ValueError(f"Collected {field}={value} is outside the validation range.")


def collect_official_rows() -> dict:
    """Collect every complete official month newer than the local dataset."""
    if not DATASET_PATH.exists():
        raise FileNotFoundError(f"Dataset not found: {DATASET_PATH}")

    existing = pd.read_csv(DATASET_PATH)
    existing_dates = pd.to_datetime(existing["month"], format="%b-%y")
    latest_local = _month_start(existing_dates.max())

    session = _request_session()
    spreadsheet_links = _fetch_monthly_spreadsheet_links(session)
    series = {
        field: _extract_workbook_series(_download_workbook(session, url), field)
        for field, url in spreadsheet_links.items()
    }

    common_months = set.intersection(*(set(values) for values in series.values()))
    candidate_months = sorted(month for month in common_months if month > latest_local)
    if not candidate_months:
        return {
            "status": "up_to_date",
            "latest_local_month": latest_local.strftime("%b-%y"),
            "latest_official_common_month": max(common_months).strftime("%b-%y"),
            "rows": [],
            "sources": {
                **spreadsheet_links,
                "usd_tzs_rate": BOT_EXCHANGE_URL,
            },
        }

    exchange_series = _fetch_exchange_rate_series(
        session,
        candidate_months[0],
        candidate_months[-1],
    )

    complete_months = [month for month in candidate_months if month in exchange_series]
    missing_exchange = [month.strftime("%b-%y") for month in candidate_months if month not in exchange_series]
    if missing_exchange:
        raise RuntimeError(
            "BOT exchange rates were missing for complete source months: "
            + ", ".join(missing_exchange)
        )

    rows = []
    for month in complete_months:
        row = {
            "month": month.strftime("%b-%y"),
            "inflation_pct": series["inflation_pct"][month],
            "usd_tzs_rate": exchange_series[month],
            "brent_oil_usd": series["brent_oil_usd"][month],
            "m2_bn_tzs": series["m2_bn_tzs"][month],
            "bot_policy_rate_pct": series["bot_policy_rate_pct"][month],
        }
        _validate_row(row)
        rows.append(row)

    sources = {
        **spreadsheet_links,
        "usd_tzs_rate": BOT_EXCHANGE_URL,
    }
    return {
        "status": "collected",
        "latest_local_month": latest_local.strftime("%b-%y"),
        "latest_official_common_month": complete_months[-1].strftime("%b-%y"),
        "rows": rows,
        "sources": sources,
    }


def update_dataset_and_models() -> dict:
    """Collect official rows, append them safely, retrain, and predict."""
    result = collect_official_rows()
    rows = result["rows"]

    if rows:
        existing = pd.read_csv(DATASET_PATH)
        additions = pd.DataFrame(rows)
        combined = pd.concat([existing, additions], ignore_index=True)
        combined["_date"] = pd.to_datetime(combined["month"], format="%b-%y")
        combined = (
            combined.drop_duplicates(subset=["_date"], keep="last")
            .sort_values("_date")
            .drop(columns="_date")
        )
        combined.to_csv(DATASET_PATH, index=False, float_format="%.6f")

    from export_models import main as export_models

    export_models()

    from predict_realtime import generate_predictions

    generate_predictions(write_json=True)

    audit = {
        "collected_at": pd.Timestamp.now(tz="Africa/Dar_es_Salaam").isoformat(),
        "status": result["status"],
        "rows_added": len(rows),
        "months_added": [row["month"] for row in rows],
        "models_retrained": True,
        "prediction_refreshed": True,
        "latest_official_common_month": result["latest_official_common_month"],
        "sources": result["sources"],
    }
    SOURCE_LOG_PATH.write_text(json.dumps(audit, indent=2))
    return audit


def source_status() -> dict:
    if not SOURCE_LOG_PATH.exists():
        return {
            "status": "not_run",
            "message": "Official data collection has not been run yet.",
            "sources": {
                **{field: BOT_MONTHLY_STATS_URL for field in SOURCE_TITLES},
                "usd_tzs_rate": BOT_EXCHANGE_URL,
            },
        }
    return json.loads(SOURCE_LOG_PATH.read_text())


if __name__ == "__main__":
    print(json.dumps(update_dataset_and_models(), indent=2))
